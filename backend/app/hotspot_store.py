"""Hotspot recommendations from ``英文字段.xlsx`` (top 5 rows per sheet).

List + detail fields follow repo root ``信息匹配.md`` / ``field_schema.py``.
Policy sheets use the same projection style as SceneResults.
"""

from __future__ import annotations

import os
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.field_schema import (
    ACHIEVEMENT_DETAIL_FIELDS,
    ACHIEVEMENT_LIST_FIELDS,
    DEMAND_DETAIL_FIELDS,
    DEMAND_LIST_FIELDS,
    ENTERPRISE_DETAIL_FIELDS,
    ENTERPRISE_LIST_FIELDS,
    EQUIPMENT_DETAIL_FIELDS,
    EQUIPMENT_LIST_FIELDS,
    EXPERT_DETAIL_FIELDS,
    EXPERT_LIST_FIELDS,
    FieldDef,
    PILOT_TEST_DETAIL_FIELDS,
    PILOT_TEST_LIST_FIELDS,
    POC_CENTER_DETAIL_FIELDS,
    POC_CENTER_LIST_FIELDS,
    PUBLIC_SERVICE_DETAIL_FIELDS,
    PUBLIC_SERVICE_LIST_FIELDS,
)

# Project root: backend/app/hotspot_store.py → parents[2]
# Local: <repo>/backend/app/... → <repo>
# Docker: /workspace/backend/app/... → /workspace
_REPO_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_XLSX = _REPO_ROOT / "英文字段.xlsx"
_FALLBACK_XLSX = _REPO_ROOT / "outputs" / "english_headers" / "英文字段.xlsx"
# Docker bind / 显式配置：HOTSPOT_XLSX=/path/to/file.xlsx
_ENV_XLSX = Path(os.environ["HOTSPOT_XLSX"]) if os.environ.get("HOTSPOT_XLSX") else None

TOP_N = 5

# Policy fields (aligned with SceneResults list + detail)
PROVINCIAL_POLICY_LIST_FIELDS: tuple[FieldDef, ...] = (
    FieldDef("item_name", "事项名称"),
    FieldDef("level", "级别"),
    FieldDef("funding_amount", "资助金额"),
)
PROVINCIAL_POLICY_DETAIL_FIELDS: tuple[FieldDef, ...] = (
    FieldDef("item_category", "事项类别"),
    FieldDef("item_category_description", "事项类别介绍"),
    FieldDef("project_name", "项目名称"),
    FieldDef("project_description", "项目介绍"),
    FieldDef("application_requirements", "申报要求"),
    FieldDef("application_channel", "申报途径"),
    FieldDef("application_url", "申报网址"),
    FieldDef("related_policy_document_name", "相关政策文件名称"),
    FieldDef("organizing_organization", "组织单位"),
)
MUNICIPAL_POLICY_LIST_FIELDS: tuple[FieldDef, ...] = (
    FieldDef("policy_category", "政策类别"),
    FieldDef("supported_region", "支持区域"),
    FieldDef("supported_entities", "支持对象"),
    FieldDef("support_content", "支持内容"),
)
MUNICIPAL_POLICY_DETAIL_FIELDS: tuple[FieldDef, ...] = (
    FieldDef("source_document", "来源文件"),
)

# sheet → (section key, label, list fields, detail fields, title key)
_SHEET_DEFS: tuple[
    tuple[str, str, str, tuple[FieldDef, ...], tuple[FieldDef, ...], str], ...
] = (
    (
        "achievements",
        "achievements",
        "成果",
        ACHIEVEMENT_LIST_FIELDS,
        ACHIEVEMENT_DETAIL_FIELDS,
        "achievement_name",
    ),
    (
        "requirements",
        "requirements",
        "需求",
        DEMAND_LIST_FIELDS,
        DEMAND_DETAIL_FIELDS,
        "requirement_name",
    ),
    (
        "expert_team",
        "expert_team",
        "专家",
        EXPERT_LIST_FIELDS,
        EXPERT_DETAIL_FIELDS,
        "expert_team_name",
    ),
    (
        "enterprises",
        "enterprises",
        "企业",
        ENTERPRISE_LIST_FIELDS,
        ENTERPRISE_DETAIL_FIELDS,
        "company_name",
    ),
    (
        "poc_center",
        "poc_center",
        "概念验证中心",
        POC_CENTER_LIST_FIELDS,
        POC_CENTER_DETAIL_FIELDS,
        "center_name",
    ),
    (
        "pilot_test_platform",
        "pilot_test_platform",
        "中试平台",
        PILOT_TEST_LIST_FIELDS,
        PILOT_TEST_DETAIL_FIELDS,
        "platform_name",
    ),
    (
        "large_scale_equipment",
        "large_scale_equipment",
        "大型仪器",
        EQUIPMENT_LIST_FIELDS,
        EQUIPMENT_DETAIL_FIELDS,
        "equipment_name",
    ),
    (
        "public_service_platform",
        "public_service_platform",
        "公共服务平台",
        PUBLIC_SERVICE_LIST_FIELDS,
        PUBLIC_SERVICE_DETAIL_FIELDS,
        "platform_name_required",
    ),
    (
        "provincial_policies",
        "provincial_policies",
        "省级政策",
        PROVINCIAL_POLICY_LIST_FIELDS,
        PROVINCIAL_POLICY_DETAIL_FIELDS,
        "item_name",
    ),
    (
        "municipal_policies",
        "municipal_policies",
        "市级政策",
        MUNICIPAL_POLICY_LIST_FIELDS,
        MUNICIPAL_POLICY_DETAIL_FIELDS,
        "policy_category",
    ),
)


def _resolve_xlsx_path() -> Path:
    if _ENV_XLSX is not None and _ENV_XLSX.is_file():
        return _ENV_XLSX
    if _DEFAULT_XLSX.is_file():
        return _DEFAULT_XLSX
    if _FALLBACK_XLSX.is_file():
        return _FALLBACK_XLSX
    raise FileNotFoundError(
        f"未找到热点数据文件：{_DEFAULT_XLSX} 或 {_FALLBACK_XLSX}"
        + (f" 或 HOTSPOT_XLSX={_ENV_XLSX}" if _ENV_XLSX else "")
    )


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\u200c", "").strip()
    return " ".join(text.split())


def _fields_to_dicts(fields: tuple[FieldDef, ...]) -> list[dict[str, str]]:
    return [{"key": f.key, "label": f.label} for f in fields]


def _merge_field_keys(
    list_fields: tuple[FieldDef, ...],
    detail_fields: tuple[FieldDef, ...],
) -> list[str]:
    seen: set[str] = set()
    keys: list[str] = []
    for f in (*list_fields, *detail_fields):
        if f.key not in seen:
            seen.add(f.key)
            keys.append(f.key)
    return keys


def _read_sheet_rows(
    workbook_path: Path,
    sheet_name: str,
    field_keys: list[str],
    limit: int,
) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        index_by_key = {name: i for i, name in enumerate(headers) if name}
        items: list[dict[str, str]] = []
        for raw in rows_iter:
            if raw is None:
                continue
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            projected: dict[str, str] = {}
            has_any = False
            for key in field_keys:
                idx = index_by_key.get(key)
                if idx is None or idx >= len(raw):
                    projected[key] = ""
                    continue
                text = _cell_to_str(raw[idx])
                projected[key] = text
                if text:
                    has_any = True
            if not has_any:
                continue
            items.append(projected)
            if len(items) >= limit:
                break
        return items
    finally:
        wb.close()


@lru_cache(maxsize=1)
def load_hotspots() -> dict[str, Any]:
    """Load and cache projected hotspot sections (top 5 per sheet)."""

    path = _resolve_xlsx_path()
    sections: list[dict[str, Any]] = []

    for (
        sheet_name,
        section_key,
        label,
        list_fields,
        detail_fields,
        title_key,
    ) in _SHEET_DEFS:
        field_keys = _merge_field_keys(list_fields, detail_fields)
        items = _read_sheet_rows(path, sheet_name, field_keys, TOP_N)
        if not items:
            continue
        sections.append(
            {
                "key": section_key,
                "label": label,
                "titleKey": title_key,
                "fields": _fields_to_dicts(list_fields),
                "detailFields": _fields_to_dicts(detail_fields),
                "items": items,
            }
        )

    return {
        "source": path.name,
        "topN": TOP_N,
        "sections": sections,
    }


def get_hotspots(*, reload: bool = False) -> dict[str, Any]:
    """Return hotspot payload; ``reload`` clears the process cache."""

    if reload:
        load_hotspots.cache_clear()
    return load_hotspots()
