"""Admin-editable sensitive-word lexicon (SQLite), seeded from Excel."""

from __future__ import annotations

import os
import sqlite3
import threading
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_DB_PATH = Path(__file__).resolve().parent.parent / "data" / "conversations.db"
_LOCK = threading.Lock()
_INITIALIZED = False

# Project root: backend/app/sensitive_word_store.py → parents[2]
_REPO_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_XLSX = _REPO_ROOT / "sensitiveword.xlsx"
_ENV_XLSX = (
    Path(os.environ["SENSITIVE_WORD_XLSX"])
    if os.environ.get("SENSITIVE_WORD_XLSX")
    else None
)

_WORDS_CACHE: list[str] | None = None


def _connect() -> sqlite3.Connection:
    _DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(_DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def _resolve_xlsx_path() -> Path | None:
    if _ENV_XLSX is not None and _ENV_XLSX.is_file():
        return _ENV_XLSX
    if _DEFAULT_XLSX.is_file():
        return _DEFAULT_XLSX
    return None


def _new_id() -> str:
    return uuid.uuid4().hex


def _invalidate_words_cache() -> None:
    global _WORDS_CACHE
    _WORDS_CACHE = None


def _row_to_item(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "word": row["word"],
        "category": row["category"] or "",
        "subcategory": row["subcategory"] or "",
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def _normalize_word(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _normalize_optional(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _import_from_excel(conn: sqlite3.Connection) -> int:
    path = _resolve_xlsx_path()
    if path is None:
        return 0

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return 0

        header_map = {
            str(cell).strip(): idx
            for idx, cell in enumerate(header)
            if cell is not None and str(cell).strip()
        }
        word_idx = header_map.get("敏感词")
        cat_idx = header_map.get("大类")
        sub_idx = header_map.get("小类")
        if word_idx is None:
            return 0

        seen: set[str] = set()
        batch: list[tuple[str, str, str, str]] = []
        for row in rows_iter:
            if not row or word_idx >= len(row):
                continue
            word = _normalize_word(row[word_idx])
            if not word:
                continue
            key = word.casefold()
            if key in seen:
                continue
            seen.add(key)
            category = (
                _normalize_optional(row[cat_idx])
                if cat_idx is not None and cat_idx < len(row)
                else ""
            )
            subcategory = (
                _normalize_optional(row[sub_idx])
                if sub_idx is not None and sub_idx < len(row)
                else ""
            )
            batch.append((_new_id(), word, category, subcategory))

        if not batch:
            return 0

        conn.executemany(
            """
            INSERT OR IGNORE INTO sensitive_words
                (id, word, category, subcategory, created_at, updated_at)
            VALUES (?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            batch,
        )
        return len(batch)
    finally:
        wb.close()


def init_sensitive_words() -> None:
    """Create table and seed from Excel when empty. Safe to call multiple times."""

    global _INITIALIZED
    with _LOCK:
        if _INITIALIZED:
            return
        conn = _connect()
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS sensitive_words (
                    id TEXT PRIMARY KEY,
                    word TEXT NOT NULL COLLATE NOCASE,
                    category TEXT NOT NULL DEFAULT '',
                    subcategory TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_sensitive_words_word
                ON sensitive_words(word)
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_sensitive_words_category
                ON sensitive_words(category)
                """
            )
            count_row = conn.execute(
                "SELECT COUNT(*) AS c FROM sensitive_words"
            ).fetchone()
            count = int(count_row["c"]) if count_row is not None else 0
            if count == 0:
                _import_from_excel(conn)
            conn.commit()
        finally:
            conn.close()
        _invalidate_words_cache()
        _INITIALIZED = True


def list_sensitive_words(
    q: str = "",
    category: str = "",
    page: int = 1,
    page_size: int = 20,
) -> dict[str, Any]:
    """Paginated admin list with optional keyword / category filters."""

    init_sensitive_words()
    query = (q or "").strip()
    cat = (category or "").strip()
    page = max(1, int(page or 1))
    page_size = max(1, min(100, int(page_size or 20)))

    where_parts: list[str] = []
    params: list[Any] = []
    if query:
        like = f"%{query}%"
        where_parts.append(
            "(word LIKE ? OR category LIKE ? OR subcategory LIKE ?)"
        )
        params.extend([like, like, like])
    if cat:
        where_parts.append("category = ?")
        params.append(cat)

    where_sql = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    with _LOCK:
        conn = _connect()
        try:
            total_row = conn.execute(
                f"SELECT COUNT(*) AS c FROM sensitive_words {where_sql}",
                params,
            ).fetchone()
            total = int(total_row["c"]) if total_row is not None else 0
            offset = (page - 1) * page_size
            rows = conn.execute(
                f"""
                SELECT id, word, category, subcategory, created_at, updated_at
                FROM sensitive_words
                {where_sql}
                ORDER BY updated_at DESC, word ASC
                LIMIT ? OFFSET ?
                """,
                [*params, page_size, offset],
            ).fetchall()
            items = [_row_to_item(row) for row in rows]
        finally:
            conn.close()

    return {
        "items": items,
        "total": total,
        "page": page,
        "pageSize": page_size,
    }


def list_categories() -> list[str]:
    """Distinct categories for admin filter UI."""

    init_sensitive_words()
    with _LOCK:
        conn = _connect()
        try:
            rows = conn.execute(
                """
                SELECT DISTINCT category
                FROM sensitive_words
                WHERE category != ''
                ORDER BY category ASC
                """
            ).fetchall()
            return [str(row["category"]) for row in rows if row["category"]]
        finally:
            conn.close()


def create_sensitive_word(body: dict[str, Any]) -> dict[str, Any]:
    """Insert a new sensitive word."""

    if not isinstance(body, dict):
        raise ValueError("请求体必须是 JSON 对象")

    word = _normalize_word(body.get("word"))
    if not word:
        raise ValueError("敏感词不能为空")
    category = _normalize_optional(body.get("category"))
    subcategory = _normalize_optional(body.get("subcategory"))
    item_id = _new_id()

    init_sensitive_words()
    with _LOCK:
        conn = _connect()
        try:
            try:
                conn.execute(
                    """
                    INSERT INTO sensitive_words
                        (id, word, category, subcategory, created_at, updated_at)
                    VALUES (?, ?, ?, ?, datetime('now'), datetime('now'))
                    """,
                    (item_id, word, category, subcategory),
                )
                conn.commit()
            except sqlite3.IntegrityError as exc:
                raise ValueError("敏感词已存在") from exc

            row = conn.execute(
                """
                SELECT id, word, category, subcategory, created_at, updated_at
                FROM sensitive_words WHERE id = ?
                """,
                (item_id,),
            ).fetchone()
            if row is None:
                raise ValueError("创建失败")
            _invalidate_words_cache()
            return _row_to_item(row)
        finally:
            conn.close()


def update_sensitive_word(item_id: str, body: dict[str, Any]) -> dict[str, Any]:
    """Update an existing sensitive word."""

    if not isinstance(body, dict):
        raise ValueError("请求体必须是 JSON 对象")

    sid = str(item_id or "").strip()
    if not sid:
        raise ValueError("缺少敏感词 id")

    word = _normalize_word(body.get("word"))
    if not word:
        raise ValueError("敏感词不能为空")
    category = _normalize_optional(body.get("category"))
    subcategory = _normalize_optional(body.get("subcategory"))

    init_sensitive_words()
    with _LOCK:
        conn = _connect()
        try:
            exists = conn.execute(
                "SELECT id FROM sensitive_words WHERE id = ?",
                (sid,),
            ).fetchone()
            if exists is None:
                raise ValueError("敏感词不存在")

            try:
                conn.execute(
                    """
                    UPDATE sensitive_words
                    SET word = ?, category = ?, subcategory = ?,
                        updated_at = datetime('now')
                    WHERE id = ?
                    """,
                    (word, category, subcategory, sid),
                )
                conn.commit()
            except sqlite3.IntegrityError as exc:
                raise ValueError("敏感词已存在") from exc

            row = conn.execute(
                """
                SELECT id, word, category, subcategory, created_at, updated_at
                FROM sensitive_words WHERE id = ?
                """,
                (sid,),
            ).fetchone()
            if row is None:
                raise ValueError("更新失败")
            _invalidate_words_cache()
            return _row_to_item(row)
        finally:
            conn.close()


def delete_sensitive_word(item_id: str) -> dict[str, Any]:
    """Delete a sensitive word by id."""

    sid = str(item_id or "").strip()
    if not sid:
        raise ValueError("缺少敏感词 id")

    init_sensitive_words()
    with _LOCK:
        conn = _connect()
        try:
            row = conn.execute(
                """
                SELECT id, word, category, subcategory, created_at, updated_at
                FROM sensitive_words WHERE id = ?
                """,
                (sid,),
            ).fetchone()
            if row is None:
                raise ValueError("敏感词不存在")
            item = _row_to_item(row)
            conn.execute("DELETE FROM sensitive_words WHERE id = ?", (sid,))
            conn.commit()
            _invalidate_words_cache()
            return item
        finally:
            conn.close()


def list_active_words() -> list[str]:
    """All words for client-side matching (cached), longest first."""

    global _WORDS_CACHE
    init_sensitive_words()
    with _LOCK:
        if _WORDS_CACHE is not None:
            return list(_WORDS_CACHE)
        conn = _connect()
        try:
            rows = conn.execute(
                "SELECT word FROM sensitive_words ORDER BY length(word) DESC, word ASC"
            ).fetchall()
            words = [str(row["word"]) for row in rows if row["word"]]
            _WORDS_CACHE = words
            return list(words)
        finally:
            conn.close()
