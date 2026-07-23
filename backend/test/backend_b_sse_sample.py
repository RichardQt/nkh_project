"""后端 B 样例：SSE 接口。

调用方（后端 A / Postman）只需传一个问题，例如::

    {"query": "这项新材料成果适合进入哪些应用场景？"}

后端 B 会：
1. 先推送业务数据（event: business）——读取同目录 test.xlsx
2. 再流式推送大模型回答（event: delta）——配置复用 backend/.env
3. 最后 event: done 结束

启动::

    cd backend
    python test/backend_b_sse_sample.py

默认监听 0.0.0.0:8001（本机与局域网均可访问）。
本机：http://127.0.0.1:8001
局域网：http://<本机局域网IP>:8001
"""

from __future__ import annotations

import json
import os
import time
import uuid
from collections.abc import AsyncIterator
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from pathlib import Path
from typing import Any

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

_BACKEND_ROOT = Path(__file__).resolve().parent.parent
_TEST_DIR = Path(__file__).resolve().parent
_BUSINESS_XLSX = _TEST_DIR / "test.xlsx"

load_dotenv(_BACKEND_ROOT / ".env")
load_dotenv()

LLM_BASE_URL = os.getenv("LLM_BASE_URL", "http://101.226.11.38:25000/v1").rstrip("/")
LLM_API_KEY = os.getenv("LLM_API_KEY", "")
LLM_MODEL = os.getenv("LLM_MODEL", "Qwen/Qwen3.6-35B-A3B")

app = FastAPI(
    title="Backend B Sample — SSE",
    version="0.3.0",
    description="只传一个问题；业务数据读 test.xlsx，再流式返回大模型结果。",
    docs_url="/docs",
    redoc_url=None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "Accept"],
)


class TaskStreamRequest(BaseModel):
    """请求体：只需要一个问题。"""

    query: str = Field(..., min_length=1, max_length=4000, description="用户问题")


def _sse(event: str, payload: dict[str, Any]) -> str:
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return f"event: {event}\ndata: {data}\n\n"


def _extract_delta_text(payload: dict[str, Any]) -> str:
    choices = payload.get("choices")
    if not isinstance(choices, list) or not choices:
        return ""
    first = choices[0]
    if not isinstance(first, dict):
        return ""

    delta = first.get("delta")
    if isinstance(delta, dict):
        content = delta.get("content")
        if isinstance(content, str) and content:
            return content

    message = first.get("message")
    if isinstance(message, dict):
        content = message.get("content")
        if isinstance(content, str) and content:
            return content

    text = first.get("text")
    return text if isinstance(text, str) else ""


def _json_safe(value: Any) -> Any:
    """把 Excel 单元格值转成可 JSON 序列化的类型。"""

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _read_business_from_xlsx(path: Path) -> list[dict[str, Any]]:
    """读取 test.xlsx：首行为表头，其余行为业务记录。"""

    if not path.is_file():
        raise FileNotFoundError(f"业务数据文件不存在：{path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers: list[str] = []
        for idx, cell in enumerate(header_row):
            if cell is None or str(cell).strip() == "":
                headers.append(f"col_{idx + 1}")
            else:
                headers.append(str(cell).strip())

        # 去掉表头末尾连续空列（openpyxl 可能扩到很宽）
        while headers and headers[-1].startswith("col_"):
            # 仅当该列在表头本身为空时才裁掉
            last_idx = len(headers) - 1
            if header_row[last_idx] is None or str(header_row[last_idx]).strip() == "":
                headers.pop()
            else:
                break

        items: list[dict[str, Any]] = []
        for row in rows:
            if row is None:
                continue
            # 整行空则跳过
            if all(cell is None or str(cell).strip() == "" for cell in row[: len(headers)]):
                continue
            record: dict[str, Any] = {}
            for col_idx, key in enumerate(headers):
                raw = row[col_idx] if col_idx < len(row) else None
                record[key] = _json_safe(raw)
            items.append(record)
        return items
    finally:
        workbook.close()


def _load_business_data(query: str, request_id: str) -> dict[str, Any]:
    """业务数据：直接返回同目录 test.xlsx 内容。"""

    try:
        items = _read_business_from_xlsx(_BUSINESS_XLSX)
        return {
            "requestId": request_id,
            "taskId": f"task_{uuid.uuid4().hex[:12]}",
            "bizCode": "BIZ_OK",
            "bizMessage": "业务数据读取成功（来源 test.xlsx）",
            "query": query,
            "source": {
                "type": "xlsx",
                "file": _BUSINESS_XLSX.name,
                "path": str(_BUSINESS_XLSX),
            },
            "total": len(items),
            "items": items,
        }
    except FileNotFoundError as exc:
        return {
            "requestId": request_id,
            "taskId": f"task_{uuid.uuid4().hex[:12]}",
            "bizCode": "BIZ_FILE_MISSING",
            "bizMessage": str(exc),
            "query": query,
            "source": {"type": "xlsx", "file": _BUSINESS_XLSX.name},
            "total": 0,
            "items": [],
        }
    except Exception as exc:  # noqa: BLE001 — 样例服务：读表失败时仍继续走 SSE
        return {
            "requestId": request_id,
            "taskId": f"task_{uuid.uuid4().hex[:12]}",
            "bizCode": "BIZ_READ_ERROR",
            "bizMessage": f"读取 test.xlsx 失败：{exc.__class__.__name__}: {exc}",
            "query": query,
            "source": {"type": "xlsx", "file": _BUSINESS_XLSX.name},
            "total": 0,
            "items": [],
        }


async def _stream_task(query: str, request: Request) -> AsyncIterator[str]:
    """meta → business(test.xlsx) → delta* → done"""

    request_id = f"req_{uuid.uuid4().hex}"
    started = time.time()

    yield _sse(
        "meta",
        {
            "requestId": request_id,
            "model": LLM_MODEL,
            "service": "backend-b-sample",
            "ts": int(time.time() * 1000),
        },
    )

    if await request.is_disconnected():
        return

    # 1) 业务数据：读取同目录 test.xlsx，先推给后端 A / 前端
    business = _load_business_data(query, request_id)
    yield _sse("business", business)

    if await request.is_disconnected():
        return

    # 2) 大模型流式文本
    if not LLM_API_KEY:
        yield _sse(
            "delta",
            {
                "content": (
                    "模型未配置 API Key（backend/.env 的 LLM_API_KEY）。"
                    "业务数据已在 business 事件中返回（来源 test.xlsx）。"
                ),
                "requestId": request_id,
            },
        )
        yield _sse(
            "done",
            {
                "requestId": request_id,
                "finishReason": "error",
                "errorCode": "LLM_NOT_CONFIGURED",
                "elapsedMs": int((time.time() - started) * 1000),
            },
        )
        return

    # 把 Excel 业务摘要塞给模型（字段过长时截断，避免 prompt 过大）
    compact_items: list[dict[str, Any]] = []
    for row in business.get("items") or []:
        if not isinstance(row, dict):
            continue
        compact_items.append(
            {
                "序号": row.get("序号"),
                "成果名称": row.get("成果名称"),
                "技术领域": row.get("技术领域"),
                "技术领域(二级)": row.get("技术领域(二级)"),
                "南京重点发展产业领域": row.get("南京重点发展产业领域"),
                "成熟度": row.get("成熟度"),
                "权利归属机构名称": row.get("权利归属机构名称"),
                "意向金额（单位：万元）": row.get("意向金额（单位：万元）"),
                "成果简介": (str(row.get("成果简介") or "")[:200] or None),
                "是否属于“碳达峰、碳中和”领域": row.get("是否属于“碳达峰、碳中和”领域"),
            }
        )

    compact = {
        "source": business.get("source"),
        "total": business.get("total"),
        "items": compact_items,
    }
    user_content = (
        f"用户问题：{query.strip()}\n\n"
        f"业务系统从 test.xlsx 读取的成果列表（JSON）：\n"
        f"{json.dumps(compact, ensure_ascii=False, indent=2)}\n\n"
        "请基于上述业务成果数据用中文简要回答：先结论，再分点说明与下一步建议。"
    )

    url = f"{LLM_BASE_URL}/chat/completions"
    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json",
        "Accept": "text/event-stream",
    }
    llm_body = {
        "model": LLM_MODEL,
        "stream": True,
        "messages": [
            {
                "role": "system",
                "content": (
                    "你是业务+AI 协同助手。结合业务结构化结果给出清晰、可执行的中文建议。"
                    "不要编造不可核验的精确数据。"
                ),
            },
            {"role": "user", "content": user_content},
        ],
    }

    timeout = httpx.Timeout(connect=20.0, read=180.0, write=30.0, pool=20.0)
    emitted_any = False

    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            async with client.stream("POST", url, headers=headers, json=llm_body) as response:
                if response.status_code >= 400:
                    error_body = (await response.aread()).decode("utf-8", errors="replace")
                    detail = error_body[:500] if error_body else f"HTTP {response.status_code}"
                    yield _sse(
                        "delta",
                        {
                            "content": f"模型服务暂时不可用（{response.status_code}）：{detail}",
                            "requestId": request_id,
                        },
                    )
                    yield _sse(
                        "done",
                        {
                            "requestId": request_id,
                            "finishReason": "error",
                            "errorCode": "LLM_HTTP_ERROR",
                            "elapsedMs": int((time.time() - started) * 1000),
                        },
                    )
                    return

                async for line in response.aiter_lines():
                    if await request.is_disconnected():
                        return
                    if not line or line.startswith(":") or not line.startswith("data:"):
                        continue

                    data = line[5:].strip()
                    if not data or data == "[DONE]":
                        if data == "[DONE]":
                            break
                        continue

                    try:
                        payload = json.loads(data)
                    except json.JSONDecodeError:
                        continue
                    if not isinstance(payload, dict):
                        continue

                    if payload.get("error"):
                        err = payload["error"]
                        err_text = err.get("message") if isinstance(err, dict) else str(err)
                        yield _sse(
                            "delta",
                            {"content": f"模型返回错误：{err_text}", "requestId": request_id},
                        )
                        yield _sse(
                            "done",
                            {
                                "requestId": request_id,
                                "finishReason": "error",
                                "errorCode": "LLM_RESPONSE_ERROR",
                                "elapsedMs": int((time.time() - started) * 1000),
                            },
                        )
                        return

                    text = _extract_delta_text(payload)
                    if text:
                        emitted_any = True
                        yield _sse("delta", {"content": text, "requestId": request_id})

    except httpx.TimeoutException:
        yield _sse(
            "delta",
            {"content": "模型响应超时。业务数据（test.xlsx）已返回。", "requestId": request_id},
        )
        yield _sse(
            "done",
            {
                "requestId": request_id,
                "finishReason": "error",
                "errorCode": "LLM_TIMEOUT",
                "elapsedMs": int((time.time() - started) * 1000),
            },
        )
        return
    except httpx.HTTPError as exc:
        yield _sse(
            "delta",
            {
                "content": f"无法连接模型服务：{exc.__class__.__name__}。业务数据（test.xlsx）已返回。",
                "requestId": request_id,
            },
        )
        yield _sse(
            "done",
            {
                "requestId": request_id,
                "finishReason": "error",
                "errorCode": "LLM_CONNECT_ERROR",
                "elapsedMs": int((time.time() - started) * 1000),
            },
        )
        return

    if not emitted_any:
        yield _sse(
            "delta",
            {
                "content": "模型未返回可用内容。可先展示 business 中的 test.xlsx 业务数据。",
                "requestId": request_id,
            },
        )

    yield _sse(
        "done",
        {
            "requestId": request_id,
            "finishReason": "stop",
            "elapsedMs": int((time.time() - started) * 1000),
            "hasBusiness": True,
            "hasModelText": emitted_any,
        },
    )


@app.get("/api/b/health")
async def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "service": "backend-b-sample",
        "modelConfigured": bool(LLM_API_KEY and LLM_BASE_URL),
        "model": LLM_MODEL,
    }


@app.post("/api/b/task/stream")
async def task_stream(body: TaskStreamRequest, request: Request) -> StreamingResponse:
    """只需传一个问题：`{"query": "你的问题"}`。

    SSE 顺序：meta → business（test.xlsx）→ delta* → done
    """

    query = body.query.strip()
    if not query:
        raise HTTPException(status_code=422, detail="query 不能为空")

    return StreamingResponse(
        _stream_task(query, request),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _lan_ipv4_hints() -> list[str]:
    """尽量列出本机局域网 IPv4，方便联调时拼 URL。"""

    import socket

    ips: list[str] = []
    try:
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None, socket.AF_INET):
            ip = info[4][0]
            if ip and not ip.startswith("127.") and ip not in ips:
                ips.append(ip)
    except OSError:
        pass

    # 再尝试一次“出网路由”探测（不真正发包）
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
            if ip and not ip.startswith("127.") and ip not in ips:
                ips.insert(0, ip)
    except OSError:
        pass

    return ips


if __name__ == "__main__":
    import uvicorn

    # 0.0.0.0：绑定所有网卡，允许局域网其它机器访问
    host = os.getenv("BACKEND_B_HOST", "0.0.0.0")
    port = int(os.getenv("BACKEND_B_PORT", "8001"))

    print(f"[backend-b] listen on http://{host}:{port}")
    print(f"[backend-b] local  : http://127.0.0.1:{port}")
    for ip in _lan_ipv4_hints():
        print(f"[backend-b] lan    : http://{ip}:{port}")
    print(f"[backend-b] docs   : http://127.0.0.1:{port}/docs")
    print(f"[backend-b] health : http://127.0.0.1:{port}/api/b/health")

    uvicorn.run(app, host=host, port=port, log_level="info")
